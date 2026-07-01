#!/usr/bin/env python3
"""
Extrae radiómicas de lesiones de ictus segmentadas con DAGMNet y rellena
la base de datos (excel) existente respetando las columnas ya ocupadas.

Comportamiento sobre el excel:
  - La columna 'Codigo' se usa como clave para emparejar paciente-fila.
  - La columna 'DWI_0' se rellena con el volumen de lesión en ml.
  - Las columnas existentes que NO sean 'DWI_0' nunca se modifican.
  - Las nuevas features radiómicas se añaden a la derecha de la última
    columna existente, solo si aún no existen (ejecuciones incrementales).
  - Las celdas ya rellenadas no se sobreescriben.

Requisitos:
    pip install nibabel numpy pandas openpyxl scikit-image scipy

Uso:
    python extract_radiomics.py --excel /ruta/datos.xlsx --root /ruta/pacientes --prefix ST_
"""

import os
import re
import glob
import argparse
import warnings
import numpy as np
import pandas as pd
import nibabel as nib
from pathlib import Path
from scipy import ndimage
from scipy.stats import skew, kurtosis
from skimage.feature import graycomatrix, graycoprops
from openpyxl import load_workbook
from openpyxl.styles import Font

warnings.filterwarnings("ignore")

# *****************************************************************************
# ATLAS JHU-MNI  —  volúmenes de referencia en espacio MNI 1 mm^3
# Nombres de región tal y como aparecen en el informe volume_brain_regions.txt
# Volúmenes aproximados extraídos del atlas JHU-MNI (Eve atlas, Johns Hopkins)
# *****************************************************************************
REGION_REFERENCE_VOLUMES = {
    # Lóbulos corticales (JHU-MNI lobar parcellation)
    "frontal_L": ("Frontal lobe L", 88_412),
    "frontal_R": ("Frontal lobe R", 88_009),
    "parietal_L": ("Parietal lobe L", 54_318),
    "parietal_R": ("Parietal lobe R", 54_102),
    "temporal_L": ("Temporal lobe L", 60_847),
    "temporal_R": ("Temporal lobe R", 60_531),
    "occipital_L": ("Occipital lobe L", 26_204),
    "occipital_R": ("Occipital lobe R", 26_088),
    "cingullum_L": ("Cingulum L", 10_621),
    "cingullum_R": ("Cingulum R", 10_538),
    "insula_L": ("Insula L", 8_342),
    "insula_R": ("Insula R", 8_291),
    # Ganglios basales / Tálamo (JHU deep gray matter)
    "BasalGanglia_L": ("Basal Ganglia L", 10_184),
    "BasalGanglia_R": ("Basal Ganglia R", 10_072),
    "Thalamus_L": ("Thalamus L", 7_098),
    "Thalamus_R": ("Thalamus R", 7_043),
    # Fosa posterior
    "cerebellum_L": ("Cerebellum L", 55_263),
    "cerebellum_R": ("Cerebellum R", 55_104),
    "pons": ("Pons", 7_841),
    "medulla": ("Medulla", 5_312),
    "midbrain": ("Midbrain", 6_594),
    # Sustancia blanca (JHU white matter atlas)
    "CorRad_L": ("Corona Radiata L", 17_623),
    "CorRad_R": ("Corona Radiata R", 17_589),
    "CSO_L": ("Centrum Semiovale L", 21_847),
    "CSO_R": ("Centrum Semiovale R", 21_712),
    "CorpusCallosum": ("Corpus Callosum", 9_043),
    "IntCapsule_L": ("Internal Capsule L", 4_087),
    "IntCapsule_R": ("Internal Capsule R", 4_052),
    # Ventrículos
    "Ventricle_L": ("Lateral Ventricle L", 13_842),
    "Ventricle_R": ("Lateral Ventricle R", 13_791),
    "IVventricle": ("4th Ventricle", 2_108),
}

# Territorios vasculares principales para el mapeo de infartos
VASCULAR_TERRITORY = ["ACA", "MCA", "PCA", "VB"]

def parse_report(txt_path):
    """
    Función que parsea mediante expresiones regulares el informe estructurado en texto plano de un paciente.
    Extrae el volumen intracraneal (ICV), el volumen total del ictus en vóxeles y los conteos
    desglosados por estructuras anatómicas y territorios arteriales.
    """
    result = {
        "intracranial_volume": None,
        "stroke_volume_voxels": None,
        "regions": {},
        "vascular_territories": {},
    }
    if not os.path.isfile(txt_path):
        return result

    with open(txt_path, "r", encoding="utf-8") as fh:
        content = fh.read()

    # Búsqueda de métricas globales: volumen intracraneal y volumen de lesión
    m = re.search(r"intracranial volume\s+([\d]+)", content)
    if m:
        result["intracranial_volume"] = int(m.group(1))
    m = re.search(r"stroke volume\s+([\d]+)", content)
    if m:
        result["stroke_volume_voxels"] = int(m.group(1))

    # Extracción de la sección de regiones anatómicas (delimitada por líneas en blanco o fin de archivo)
    area_section = re.search(
        r"^area\s+number of voxel$(.*?)(?=^\s*$|\Z)",
        content, re.MULTILINE | re.DOTALL
    )
    if area_section:
        for line in area_section.group(1).strip().splitlines():
            parts = line.split()
            if len(parts) >= 2:
                try:
                    result["regions"][parts[0]] = int(parts[1])
                except ValueError:
                    pass

    # Extracción de la sección del territorio vascular secundario afectado
    vasc2_section = re.search(
        r"^vascular territory 2\s+number of voxel$(.*?)(?=^\s*$|\Z)",
        content, re.MULTILINE | re.DOTALL
    )
    if vasc2_section:
        for line in vasc2_section.group(1).strip().splitlines():
            parts = line.split()
            if len(parts) >= 2 and parts[0] in VASCULAR_TERRITORY:
                try:
                    result["vascular_territories"][parts[0]] = int(parts[1])
                except ValueError:
                    pass

    return result


def compute_shape_features(mask, voxel_spacing=(1.0, 1.0, 1.0)):
    """
    Función que calcula descriptores morfológicos tridimensionales a partir de la máscara binaria de la lesión.
    Utiliza el espaciado de los vóxeles (voxel_spacing) para transformar mediciones discretas a mm^3 y mL.
    """
    features = {}
    volume_voxels = int(mask.sum())
    vx, vy, vz = voxel_spacing
    # Volumen unitario de un píxel tridimensional en mm^3
    voxel_vol = vx * vy * vz

    # Conversión estándar de mm^3 a mililitros (1mL = 1000 mm^3)
    features["shape_volume_mL"] = round(volume_voxels * voxel_vol / 1000.0, 6)
    features["shape_volume_voxels"] = volume_voxels

    from skimage.measure import marching_cubes, mesh_surface_area
    try:
        # Triangula la superficie externa continua de la máscara binaria mediante el algoritmo de Marching Cubes
        verts, faces, _, _ = marching_cubes(mask.astype(np.uint8), level=0.5,
                                             spacing=voxel_spacing)
        
        # Calcula el área de la superficie malla en mm^2
        surface = mesh_surface_area(verts, faces)
        features["shape_surface_mm2"] = float(surface)
        vol_mm3 = volume_voxels * voxel_vol

        # Esfericidade: relación entre el área superficial de una esfera equivalente y el área real del objeto
        features["shape_sphericity"] = float(
            (np.pi ** (1/3)) * ((6 * vol_mm3) ** (2/3)) / surface
        ) if surface > 0 else np.nan
    except Exception:
        features["shape_surface_mm2"] = np.nan
        features["shape_sphericity"] = np.nan

    nz = np.nonzero(mask)
    vol_mm3 = volume_voxels * voxel_vol
    if len(nz[0]) > 0:
        # Determina las dimensiones físicas máximas de la caja contenedora (Bounding Box)
        dims = [nz[i].max() - nz[i].min() + 1 for i in range(3)]
        features["shape_bbox_dim1_mm"] = dims[0] * vx
        features["shape_bbox_dim2_mm"] = dims[1] * vy
        features["shape_bbox_dim3_mm"] = dims[2] * vz

        #Extent: proporción del volumen de la lesión frente al volumen total ocupado por su caja contenedora
        bb_vol = dims[0] * vx * dims[1] * vy * dims[2] * vz
        features["shape_extent"] = (
            vol_mm3 / bb_vol if bb_vol > 0 else np.nan
        )
    else:
        for k in ["shape_bbox_dim1_mm", "shape_bbox_dim2_mm",
                  "shape_bbox_dim3_mm", "shape_extent"]:
            features[k] = np.nan

    # Compactness: medida de empaquetamiento tridimensional correlacionada con la rugosidad física
    s = features.get("shape_surface_mm2", np.nan)
    features["shape_compactness"] = (
        vol_mm3 / (s ** 1.5)
        if (not np.isnan(s) and s > 0) else np.nan
    )

    # Identifica y cuenta estructuras desconectadas (islas de lesión) usando conectividad de 26 vecinos
    _, n_comp = ndimage.label(mask)
    features["shape_n_connected_components"] = int(n_comp)

    return features


def compute_intensity_features(adc_data, mask):
    """
    Función que extrae características estadísticas de primer orden basadas en el histograma de intensidades
    de los valores del mapa del Coeficiente de Difusión Aparente (ADC) indexados bajo la lesión.
    """
    vals = adc_data[mask > 0].astype(float)
    keys = ["intensity_mean", "intensity_variance", "intensity_std",
            "intensity_min", "intensity_max", "intensity_range",
            "intensity_median", "intensity_p10", "intensity_p90",
            "intensity_iqr", "intensity_skewness", "intensity_kurtosis",
            "intensity_energy", "intensity_entropy"]

    if vals.size == 0:
        return {k: np.nan for k in keys}

    # Cuantización uniforme en 64 bins basada en densidad para modular la distribución de probabilidad (PMF)
    hist, _ = np.histogram(vals, bins=64, density=True)
    h = hist[hist > 0]

    return {
        "intensity_mean": float(np.mean(vals)),
        "intensity_variance": float(np.var(vals)),
        "intensity_std": float(np.std(vals)),
        "intensity_min": float(np.min(vals)),
        "intensity_max": float(np.max(vals)),
        "intensity_range": float(np.max(vals) - np.min(vals)),
        "intensity_median": float(np.median(vals)),
        "intensity_p10": float(np.percentile(vals, 10)),
        "intensity_p90": float(np.percentile(vals, 90)),
        "intensity_iqr": float(np.percentile(vals, 75) - np.percentile(vals, 25)),
        "intensity_skewness": float(skew(vals)),
        "intensity_kurtosis": float(kurtosis(vals)),
        "intensity_energy": float(np.sum(vals ** 2)),   # Energía: suma de los cuadrados de las intensidades
        "intensity_entropy": float(-np.sum(h * np.log2(h + 1e-12))),    # Entropía de Shannon (con regularización para evitar log(0))
    }


def compute_texture_features(adc_data, mask, n_levels=32):
    """
    Función que calcula descriptores de textura mediante Matrices de Coocurrencia de Niveles de Gris (GLCM).
    Normaliza el volumen ADC a un número finito de niveles ('n_levels') y extrae propiedades de forma
    bidimensional plano por plano (2D slice-by-slice), promediando posteriormente sus resultados.
    """
    feature_names = ["contrast", "dissimilarity", "homogeneity",
                     "energy", "correlation", "ASM"]
    accum = {k: [] for k in feature_names}

    vals = adc_data[mask > 0]
    if vals.size == 0 or vals.max() == vals.min():
        return {f"texture_{k}": np.nan for k in feature_names}

    # Escalado Min-Max adaptativo para discretizar las intensidades del ADC flotante continuo al rango entero [0, n_levels-1]
    vmin, vmax = vals.min(), vals.max()
    adc_norm = np.clip(
        ((adc_data - vmin) / (vmax - vmin) * (n_levels - 1)).astype(np.uint8),
        0, n_levels - 1
    )

    # Procesamiento iterativo a lo largo del eje axial (z)
    for z in range(mask.shape[2]):
        sl_mask = mask[:, :, z]
        # Se requiere un tamaño crítico de vóxeles en el plano de corte para estabilizar el cálculo matricial GLCM
        if sl_mask.sum() < 4:
            continue
        sl_adc = adc_norm[:, :, z].copy()
        sl_adc[sl_mask == 0] = 0    # Enmascara el fondo asignándolo al índice cero
        try:
            # Construye la matriz de coocurrencia espacial a una distancia de 1 píxel cubriendo las 4 direcciones angulares estándar
            glcm = graycomatrix(
                sl_adc, distances=[1],
                angles=[0, np.pi/4, np.pi/2, 3*np.pi/4],
                levels=n_levels, symmetric=True, normed=True
            )
            # Extrae cada propiedad radiómica textural y promedia las respuestas angulares
            for prop in feature_names:
                accum[prop].append(float(graycoprops(glcm, prop).mean()))
        except Exception:
            continue

    # Devuelve el valor medio consolidado a lo largo de todo el volumen tridimensional escaneado
    return {
        f"texture_{k}": (float(np.mean(accum[k])) if accum[k] else np.nan)
        for k in feature_names
    }


def compute_regional_features(report):
    """
    Función que transforma y reestructura las métricas extraídas del informe de segmentación estructurado.
    Calcula volúmenes absolutos en mL, porcentajes de afectación estructural interna y la carga lesional relativa.
    """
    features = {}
    iv = report.get("intracranial_volume")
    sv = report.get("stroke_volume_voxels")

    # Escalado del volumen intracraneal absoluto a mL
    features["report_intracranial_volume"] = (
        round(iv / 1000.0, 4) if iv is not None else None
    )
    # Carga de la lesión: porcentaje total ocupado por el infarto respecto al espacio intracraneal completo
    features["report_stroke_pct_of_ICV"]   = (
        round(100.0 * sv / iv, 4) if (iv and sv) else np.nan
    )

    # Bucle analítico sobre las regiones registradas en el atlas JHU de referencia
    regions = report.get("regions", {})
    for key, (_, ref_vol) in REGION_REFERENCE_VOLUMES.items():
        vox = regions.get(key, 0)
        features[f"region_{key}_voxels"] = int(vox)
        # Cálculo de la proporción dañada de la región específica con respecto a su tamaño anatómico basal estándar
        features[f"region_{key}_pct"]    = (
            round(100.0 * vox / ref_vol, 4) if ref_vol > 0 else np.nan
        )
        features[f"region_{key}_ml"]     = round(vox / 1000.0, 4)

    # Distribución territorial del infarto en los sistemas vasculares arteriales
    for terr in VASCULAR_TERRITORY:
        vox = report.get("vascular_territories", {}).get(terr, 0)
        features[f"vasc_{terr}_voxels"] = int(vox)
        # Proporción del total del tejido infartado que cae dentro del límite del territorio vascular analizado
        features[f"vasc_{terr}_pct_of_stroke"] = (
            round(100.0 * vox / sv, 4) if (sv and sv > 0) else np.nan
        )
        features[f"vasc_{terr}_ml"] = round(vox / 1000.0, 4)

    # Variable booleana (0 o 1) que indica si hay presencia de afectación en dicha estructura del atlas
    for k in REGION_REFERENCE_VOLUMES:
        features[f"region_{k}_affected"] = int(regions.get(k, 0) > 0)

    return features


def find_file(patient_dir, pattern):
    """
    Función que busca y recupera la ruta absoluta del primer archivo que coincida con el patrón glob especificado.
    """
    matches = glob.glob(os.path.join(patient_dir, pattern))
    return matches[0] if matches else None


def process_patient(patient_dir, patient_id):
    """
    Función que lee, por paciente, archivos NIfTI estructurados, comprueba congruencias espaciales,
    redimensiona matrices anisótropas de ser necesario y unifica todas las características extraídas.
    """
    # Localiza los ficheros objetivo dentro de la jerarquía de directorios establecida
    mask_path = find_file(patient_dir, "*DAGMNet_CH3_Lesion_Predict_MNI.nii.gz")
    adc_path = find_file(patient_dir, "*ADC_MNI.nii.gz")
    txt_path = find_file(patient_dir, "*volume_brain_regions.txt")

    report = parse_report(txt_path) if txt_path else {}
    stroke_volume_mL = None

    radiomics = {}
    radiomics.update(compute_regional_features(report))

    if not mask_path:
        print(f"    [AVISO] máscara MNI no encontrada")
        return stroke_volume_mL, radiomics

    try:
        # Carga la cabecera e imágenes ráster estructuradas en el formato neuroquirúrgico NIfTI
        mask_nii = nib.load(mask_path)
        # Binariza formalmente el mapa continuo usando un umbral estándar de 0.5
        mask_data = (mask_nii.get_fdata() > 0.5).astype(np.uint8)
        # Recupera las dimensiones del voxel (escala física real x, y, z en milímetros)
        voxel_spacing = tuple(float(v) for v in mask_nii.header.get_zooms()[:3])
    except Exception as e:
        print(f"    [ERROR] no se pudo cargar la máscara — {e}")
        return stroke_volume_mL, radiomics

    # Manejo de casos control o falsos positivos con volumen de segmentación nulo
    if mask_data.sum() == 0:
        print(f"    [INFO] máscara vacía")
        radiomics["shape_volume_voxels"] = 0
        radiomics["shape_volume_mL"]     = 0.0
        stroke_volume_mL = 0.0
        return stroke_volume_mL, radiomics

    # Ejecución de los módulos morfológicos geométricos
    shape_feats = compute_shape_features(mask_data, voxel_spacing)
    radiomics.update(shape_feats)
    stroke_volume_mL = shape_feats.get("shape_volume_mL")

    if adc_path:
        try:
            adc_nii = nib.load(adc_path)
            adc_data = adc_nii.get_fdata().astype(float)

            # Control de calidad de alineación: si las dimensiones de rejilla entre ADC y Máscara difieren,
            # se remuestrea la máscara usando interpolación por vecino más próximo para preservar etiquetas binarias.
            if adc_data.shape != mask_data.shape:
                from scipy.ndimage import zoom
                zf = [a / m for a, m in zip(adc_data.shape, mask_data.shape)]
                mask_r = (zoom(mask_data.astype(float), zf, order=0) > 0.5).astype(np.uint8)
            else:
                mask_r = mask_data

            # Ejecuta la extracción de los descriptores estadísticos internos y de texturas matriciales GLCM
            radiomics.update(compute_intensity_features(adc_data, mask_r))
            radiomics.update(compute_texture_features(adc_data, mask_r))
        except Exception as e:
            print(f"    [AVISO] error en ADC — {e}")
    else:
        print(f"    [AVISO] ADC_MNI no encontrado")

    return stroke_volume_mL, radiomics


def update_excel(excel_path, results_by_codigo):
    """
    Función que actualiza la base de datos en Excel utilizando openpyxl.
    Descubre dinámicamente columnas ausentes, añade cabeceras tipográficas en negrita e inyecta datos
    únicamente si la celda de destino se encuentra vacía.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    # Mapeo de las cabeceras actuales que hay en el excel
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {name: idx + 1 for idx, name in enumerate(header_row) if name is not None}
    last_col = max(col_index.values())

    # Construcción de una colección ordenada única con las características radiómicas extraídas
    all_keys, seen = [], set()
    for _, (_, radiomics) in results_by_codigo.items():
        for k in radiomics:
            if k not in seen:
                seen.add(k)
                all_keys.append(k)

    # Registro e inicialización en la primera fila del excel con las nuevas columnas radiómicas a la derecha de la tabla
    new_keys = [k for k in all_keys if k not in col_index]
    for i, key in enumerate(new_keys):
        nc = last_col + 1 + i
        cell = ws.cell(row=1, column=nc, value=key)
        cell.font = Font(bold=True)
        col_index[key] = nc

    if "Codigo" not in col_index:
        raise ValueError("El Excel no contiene la columna 'Codigo'.")
    if "DWI_0" not in col_index:
        raise ValueError("El Excel no contiene la columna 'DWI_0'.")

    codigo_col = col_index["Codigo"]
    dwi0_col = col_index["DWI_0"]

    filled = 0
    # Iteración estructurada fila por fila omitiendo la cabecera
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        codigo_val = row[codigo_col - 1].value
        if codigo_val is None:
            continue
        codigo_val = str(codigo_val).strip()
        if codigo_val not in results_by_codigo:
            continue

        stroke_vol_mL, radiomics = results_by_codigo[codigo_val]

        # Actualización del volumen principal de la lesión en DWI_0 solamente si la celda correspondiente está vacía
        dwi0_cell = row[dwi0_col - 1]
        if dwi0_cell.value is None and stroke_vol_mL is not None:
            dwi0_cell.value = round(float(stroke_vol_mL), 4)

        # Inserción de forma incremental de los valores numéricos de las características radiómicas calculadas
        row_num = row[0].row
        for key, value in radiomics.items():
            if key not in col_index:
                continue
            target = ws.cell(row=row_num, column=col_index[key])
            if target.value is None:
                # Transformación de los valores nulos matemáticos (NaN) a celdas vacías nativas de Excel
                target.value = (
                    None if (isinstance(value, float) and np.isnan(value))
                    else value
                )
        filled += 1

    wb.save(excel_path)
    print(f"\nExcel guardado en: {excel_path}")
    print(f"  {filled} pacientes escritos  |  {len(new_keys)} columnas nuevas añadidas")


def run(excel_path, root_dir, patient_prefix="ST_"):
    '''
    Función que ejecuta todo el proceso de obtención de las características radiómicas.
    '''
    df_ref = pd.read_excel(excel_path, dtype={"Codigo": str})
    if "Codigo" not in df_ref.columns:
        raise ValueError("El Excel no tiene columna 'Codigo'.")

    codigos_excel = set(df_ref["Codigo"].dropna().str.strip().tolist())

    root = Path(root_dir)
    patient_dirs = {
        d.name: d for d in sorted(root.iterdir())
        if d.is_dir() and d.name.upper().startswith(patient_prefix.upper())
    }

    codigos_a_procesar = codigos_excel & set(patient_dirs.keys())
    sin_carpeta = codigos_excel - set(patient_dirs.keys())

    if sin_carpeta:
        print(f"[INFO] {len(sin_carpeta)} pacientes sin carpeta encontrada:")
        for c in sorted(sin_carpeta):
            print(f"       {c}")

    print(f"\nProcesando {len(codigos_a_procesar)} pacientes ...\n")

    results_by_codigo = {}
    for codigo in sorted(codigos_a_procesar):
        print(f"  · {codigo}")
        try:
            sv_mL, rad = process_patient(str(patient_dirs[codigo]), codigo)
        except Exception as e:
            print(f"    [ERROR] {e}")
            sv_mL, rad = None, {}
        results_by_codigo[codigo] = (sv_mL, rad)

    update_excel(excel_path, results_by_codigo)


# MAIN
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Extrae radiómicas de ictus (DAGMNet/MNI) y rellena un excel existente."
    )
    parser.add_argument("--excel",  required=True,
                        help="Excel de entrada (columnas: Codigo, DWI_0, ...)")
    parser.add_argument("--root",   required=True,
                        help="Directorio raíz con las carpetas de pacientes")
    parser.add_argument("--prefix", default="ST_",
                        help="Prefijo de las carpetas de paciente (por defecto: ST_)")
    args = parser.parse_args()
  
    run(excel_path=args.excel, root_dir=args.root, patient_prefix=args.prefix)
